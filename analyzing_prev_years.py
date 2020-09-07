import openpyxl
import string
from sportsreference.nfl.teams import Teams as nflTeams
import PySimpleGUI as sg
import matplotlib.pyplot as plt

#load the NFL teams
teams = nflTeams()

#load the workbooks of the data, and of the analysis place to input findings
wk_data = openpyxl.load_workbook("prev_years_data.xlsx")
wk_analysis = openpyxl.load_workbook("prev_years_analysis.xlsx")

team_results = {}
team_names = []
for team in teams:
    team_results[team.name] = []            #creates the dictionary of results for each team
    team_names.append(team.name)            #creates the list of team names
team_names = tuple(team_names)

for team in team_results:
    for year in range(13):
        team_results[team].append([0])       #creates the list of spread success for each team

status_stats = []
for year in range(13):
    status_stats.append([])
    for week in range(18):
        status_stats[year].append([0,0,0,0])       #creates the list for yearly game types[home favorite,away uunderdog,home underdog,away favorite]

for sheet in wk_data:
    year_i = 2018 - sheet.cell(1,1).value     #the index for the year, 2018:0, 2017:1, 2016:2...2006:12

    #going through the rows in the data excel sheet
    for row in range(1,102):
        if row % 6 == 1:    #if you've gone to the next week
            week = row//6 + 1
            for team in team_results:
                team_results[team][year_i].append(team_results[team][year_i][week-1])  #making a result for the new week

        for col in range(4,20):
            cell = sheet.cell(row,col).value
            if cell == None:        #when its the end of all the matches for the week
                break
            if row % 6 == 1:                           #if the team beat the spread, their spread success history goes up by one
                team_results[cell][year_i][week] += 1
            elif row % 6 == 5:                         #if the team lost to the spread their spread success history goes down by one
                team_results[cell][year_i][week] -= 1
            elif row % 6 == 2:                         #updating the count for win by the game type [hf,au,hu,af]
                if cell == 'home' and sheet.cell(row+1,col).value == 'favorite':
                    status_stats[year_i][week][0] += 1
                elif cell == 'away' and sheet.cell(row+1,col).value == 'underdog':
                    status_stats[year_i][week][1] += 1
                elif cell == 'home' and sheet.cell(row+1,col).value == 'underdog':
                    status_stats[year_i][week][2] += 1
                elif cell == 'away' and sheet.cell(row+1,col).value == 'favorite':
                    status_stats[year_i][week][3] += 1

sh_overall = wk_analysis["Overall"]

#finding the year end spread success for each team by year
year_end_sum = {}
for team in team_results:
    year_end_sum[team] = []
    for year in team_results[team]:
        year_end_sum[team].append(year[-1])

#creates a list of the years in ascending order for the graphs in the gui
gui_years = [str(val) for val in range(2018,2005,-1)]

#creates the gui of windows from PySimpleGui
sg.ChangeLookAndFeel('BlueMono')

#creating the layout for the home window
layout_h = [
        [sg.Button('Overall Statistics')],
        [sg.Button('Team Statistics by Year')],
        [sg.Exit()]
        ]

#creating the layout for the window of yearly team statistics
layout_y = [
        [sg.Text('Choose a Year'), sg.Combo(gui_years, key = 'chosen_year')],
        [sg.Text('Choose a Team to Analyze'), sg.Combo(team_names, key = 'chosen_team')],
        [sg.Submit()],
        [sg.Button('Clear Graph')],
        [sg.Exit()]
        ]

#creating the layour for the window with the overall stats
gui_stats = ['home', 'away', 'favorite', 'underdog', 'home favorite', 'home underdog', 'away favorite', 'away underdog']
layout_o = [
        [sg.Text('Choose a Stat'), sg.Combo(gui_stats, key = 'o_chosen_stat'), sg.Button('Analyze Stat')],
        [sg.Text('Choose a Team to Analyze'), sg.Combo(team_names, key = 'o_chosen_team'), sg.Button('Analyze Team'), sg.Text('Total: ', key = 'tot_team')],
        [sg.Button('Clear Graph')],
        [sg.Exit()]
        ]

#creates the window from the layout
home_window  = sg.Window("Josh\'s NFL Spread Analysis").Layout(layout_h)
yearly_window  = sg.Window("Stats By Year").Layout(layout_y)
overall_window  = sg.Window("Overall Statistics").Layout(layout_o)

while True:
    h_button, h_values = home_window.Read()         #get the button pressed and values on screen at that time

    if h_button is None or h_button == 'Exit':         #exits the program
        break

    if h_button is 'Team Statistics by Year':
        while True:
            y_button, y_values = yearly_window.Read()   #opens the window of team stats for a given year
            if y_button is None or y_button == 'Exit':
                break

            if y_button is 'Submit':
                chosen_sh = wk_analysis[y_values['chosen_year']]      #choose the sheet for the chosen year
                place_year = 2018 - int(y_values['chosen_year'])

                #create a plot of the data, showing the spread success week week
                plt.plot(range(18), team_results[y_values['chosen_team']][place_year], label = y_values['chosen_year'] + ' ' + y_values['chosen_team'])
                plt.xlabel('Week')
                plt.ylabel('Spread Success')
                plt.legend(loc="upper left")
                plt.show(block=False)

            if y_button is 'Clear Graph':           #clears the graph
                plt.cla()
                plt.show(block=False)
        yearly_window.close()

    if h_button is 'Overall Statistics':
        prev = 'Clear'
        while True:
            o_button, o_values = overall_window.Read()    #opens the overall stats window
            if o_button is None or o_button == 'Exit':
                break

            if o_button is 'Analyze Team':
                if prev != o_button:    #clears the graph if was just showing overall stat graph
                    plt.clf()
                prev = o_button
                team_overall = []
                team_sum = 0

                for year in team_results[o_values['o_chosen_team']]:
                    team_overall.insert(0,year[-1])
                    team_sum += year[-1]

                overall_window['tot_team']('Total: ' + str(team_sum))
                plt.plot(range(2006,2019), team_overall, marker='o',
                        label='year by year spread success for ' + o_values['o_chosen_team'])
                plt.xlabel('Year')
                plt.ylabel('Spread Success')
                plt.legend(loc="upper left")
                plt.show(block=False)
            if o_button is 'Analyze Stat':
                if prev != o_button:
                    plt.clf()
                prev = o_button
                stat_overall = []
                stat_plot = []
                tot_stats = {'home': 0, 'away': 0, 'favorite': 0, 'underdog': 0, 'home favorite': 0, 'home underdog': 0,
                             'away favorite': 0, 'away underdog': 0}
                for year in range(13):
                    stat_overall.append({'home':0, 'away':0, 'favorite':0, 'underdog':0, 'home favorite':0, 'home underdog':0, 'away favorite':0, 'away underdog':0})
                    for week in status_stats[12-year]:
                        stat_overall[year]['home'] += week[0]+week[2]
                        stat_overall[year]['away'] += week[1]+week[3]
                        stat_overall[year]['favorite'] += week[0]+week[3]
                        stat_overall[year]['underdog'] += week[1]+week[2]
                        stat_overall[year]['home favorite'] += week[0]
                        stat_overall[year]['home underdog'] += week[2]
                        stat_overall[year]['away favorite'] += week[3]
                        stat_overall[year]['away underdog'] += week[1]
                    for stat in stat_overall[year]:
                        tot_stats[stat] += stat_overall[year][stat]
                    stat_overall[year]['home'] = 100*stat_overall[year]['home']/(stat_overall[year]['home']+stat_overall[year]['away'])
                    stat_overall[year]['away'] = 100 - stat_overall[year]['home']
                    stat_overall[year]['favorite'] = 100 * stat_overall[year]['favorite'] / (stat_overall[year]['favorite'] + stat_overall[year]['underdog'])
                    stat_overall[year]['underdog'] = 100 - stat_overall[year]['favorite']
                    stat_overall[year]['home favorite'] = 100 * stat_overall[year]['home favorite'] / (stat_overall[year]['home favorite'] + stat_overall[year]['away underdog'])
                    stat_overall[year]['away underdog'] = 100 - stat_overall[year]['home favorite']
                    stat_overall[year]['home underdog'] = 100 * stat_overall[year]['home underdog'] / (stat_overall[year]['home underdog'] + stat_overall[year]['away favorite'])
                    stat_overall[year]['away favorite'] = 100 - stat_overall[year]['home underdog']
                for year in range(13):
                    stat_plot.append(stat_overall[year][o_values['o_chosen_stat']])
                print(tot_stats[o_values['o_chosen_stat']])
                plt.plot(range(2006, 2019), stat_plot, marker='o',
                         label='year by year spread success of ' + o_values['o_chosen_stat'])
                plt.xlabel('Year')
                plt.ylabel('Spread Success by percent')
                plt.legend(loc="upper left")
                plt.show(block=False)
            if o_button is 'Clear Graph':
                plt.cla()
                plt.show(block=False)
        overall_window.close()
home_window.close()


wk_analysis.save(("prev_years_analysis.xlsx"))


