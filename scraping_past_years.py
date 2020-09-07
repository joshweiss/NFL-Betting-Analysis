from bs4 import BeautifulSoup
import openpyxl
import requests
import random
from sportsreference.nfl.teams import Teams as nflTeams
from sportsreference.nfl.boxscore import Boxscores

#loading the NFL Teams
teams = nflTeams()

#loading the excel workbook to store the data in
wk = openpyxl.load_workbook("C:\\Users\\Joshua Weiss\\Desktop\\Football Project\\Analysis Prev Years\\prev_years_data.xlsx")

curr_row = 0

#going through each of the 17 football weeks
for week_num in range(1,18):
    source = requests.get('http://www.footballlocks.com/nfl_point_spreads_week_{a}.shtml'.format(a=week_num)).text

    soup = BeautifulSoup(source, 'lxml')

    #locating the tables containing the spread info
    table = soup.find_all('table', cellspacing = "8")

    #skipping the first unrelated table that existed for weeks 1-4
    if week_num in range(1,5):
        table = table[1:]

    year = 2018
    #going through the table elements
    for index, elem in enumerate(table):
        matches = elem.find_all('tr')[1:]    #locating the matches from the table

        if len(matches) < 5:   #checking to see if it the main bulk of matches or just the Monday night games
            continue
        if len(matches) < 16:
            matches.extend(table[index+1].find_all('tr')[1:])     #if all the games are not Sunday Night, append the Monday night games in the next table
        col = 3

        #loading the sheet for the year that the current matches belong to.
        sh = wk[str(year)]

        #adding the row titles in the sheet
        for i in range(1,6):
            sh.cell(curr_row+i,1).value = year
            sh.cell(curr_row+i,2).value = week_num
        sh.cell(curr_row+1, 3).value, sh.cell(curr_row+2, 3).value, sh.cell(curr_row+3, 3).value, sh.cell(curr_row+4, 3).value, sh.cell(curr_row+5,3).value = 'Winner', 'Home/Away', 'Fav/Und', 'Spread', 'Loser'

        #loading the results from the matches of the given week in the given year
        scores = Boxscores(week_num, year).games['{}-{}'.format(week_num, year)]

        #going through each match that happened that week
        for match in matches:
            col += 1
            items = match.find_all('td')
            fav = items[1].text.split(' ')[-1]        #finding the team favored to win
            und = items[3].text.split(' ')[-1]        #finding the underdog team expected to lose

            #a series of corrections to the strings based on the unorganized nature of the source html
            if '(' in fav:
                fav = fav.split('\"')[0]
            if '(' in und:
                und = und.split('\"')[0]
            if fav == "Bay":
                fav = items[1].text.split(' ')[-2] + ' ' + items[1].text.split(' ')[-1]
            if und == "Bay":
                und = items[3].text.split(' ')[-2] + ' ' + items[3].text.split(' ')[-1]
            if 'PK' in items[2].text:
                spread = 0.0
            elif 'PPD' in items[2].text:
                spread = 2.5
            else:
                spread = abs(float(items[2].text))

            #updating the teams that have moved cities between the years to call them uniformly by their new team name
            foundf, foundu = False, False
            if "Diego" in fav: fav = "Los Angeles Chargers"; foundf = True;
            if "Diego" in und: und = "Los Angeles Chargers"; foundu = True;
            if "Louis" in fav: fav = "Los Angeles Rams"; foundf = True;
            if "Louis" in und: und = "Los Angeles Rams"; foundu = True;
            for team in teams:
                if fav in team.name and foundf == False:
                    fav = team.name
                    foundf = True
                if und in team.name and foundu == False:
                    und = team.name
                    foundu = True
                if foundf and foundu:
                    break

            #determining if the game is of the type where the home team is also the favored, or is the underdog
            if items[1].text.split(' ')[0] == 'At':
                home = fav
                away = und
                sp_type = 'home_fav'
            else:
                home = und
                away = fav
                sp_type = 'home_und'

            #randomly adding or subtracting half a point so there must be a winner and loser against the spread
            if spread.is_integer():
                spread += random.choice([-0.5, 0.5])

            #updating the excel with the collected data for the match accordingly
            for score in scores:
                if score['winning_name'] == None or score['winning_name'].split(' ')[-1] == und.split(' ')[-1]:
                    if sp_type == 'home_fav':
                        sh.cell(curr_row + 1, col).value, sh.cell(curr_row + 2, col).value, sh.cell(curr_row + 3,col).value, sh.cell(curr_row + 4, col).value, sh.cell(curr_row + 5,col).value = und, 'away', 'underdog', spread, fav
                        break
                    else:
                        sh.cell(curr_row + 1, col).value, sh.cell(curr_row + 2, col).value, sh.cell(curr_row + 3,col).value, sh.cell(curr_row + 4, col).value, sh.cell(curr_row + 5,col).value = und, 'home', 'underdog', spread, fav
                        break
                elif score['winning_name'].split(' ')[-1] == fav.split(' ')[-1]:
                    dif = float(abs(score['home_score'] - score['away_score']))
                    if dif > spread:
                        if sp_type == 'home_fav':
                            sh.cell(curr_row+1,col).value, sh.cell(curr_row+2,col).value, sh.cell(curr_row+3,col).value, sh.cell(curr_row+4,col).value, sh.cell(curr_row+5,col).value = fav, 'home', 'favorite', spread, und
                            break
                        else:
                            sh.cell(curr_row+1,col).value, sh.cell(curr_row+2,col).value, sh.cell(curr_row+3,col).value, sh.cell(curr_row+4,col).value, sh.cell(curr_row+5,col).value = fav, 'away', 'favorite', spread, und
                            break
                    else:
                        if sp_type == 'home_fav':
                            sh.cell(curr_row+1,col).value, sh.cell(curr_row+2,col).value, sh.cell(curr_row+3,col).value, sh.cell(curr_row+4,col).value, sh.cell(curr_row+5,col).value = und, 'away', 'underdog', spread, fav
                            break
                        else:
                            sh.cell(curr_row+1,col).value, sh.cell(curr_row+2,col).value, sh.cell(curr_row+3,col).value, sh.cell(curr_row+4,col).value, sh.cell(curr_row+5,col).value = und, 'home', 'underdog', spread, fav
                            break
            #if there was an issue and the excel didn't update, print the problematic placement
            if sh.cell(curr_row + 1, col).value == None:
                print(year)
                print(fav)
                print(und)
                print('--------------')
        year -= 1
    curr_row += 6

#saving the excel workbook once all the data has been placed
wk.save("C:\\Users\\Joshua Weiss\\Desktop\\Football Project\\Analysis Prev Years\\prev_years_data.xlsx")