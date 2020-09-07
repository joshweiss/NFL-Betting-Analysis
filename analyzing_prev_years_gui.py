import openpyxl
import string
from sportsreference.nfl.teams import Teams as nflTeams
import PySimpleGUI as sg
import matplotlib.pyplot as plt

wk_data = openpyxl.load_workbook("prev_years_data.xlsx")
wk_analysis = openpyxl.load_workbook("prev_years_analysis.xlsx")

#load the NFL teams
teams = nflTeams()

team_results = {}
team_names = []
for team in teams:
    print(team)
    team_results[team.name] = []            #creates the dictionary of results for each team
    team_names.append(team.name)            #creates the list of team names
team_names = tuple(team_names)

"""
for team in team_results:
    for year in range(13):
        team_results[team].append([0])       #creates the list of spread results for each team

status_stats = []
for year in range(13):
    status_stats.append([])
    for week in range(18):
        status_stats[year].append([0,0,0,0])       #creates the list for yearly [hf,au,hu,af]

for sheet in wk_data:
    #print(sheet)
    year_i = 2018 - sheet.cell(1,1).value     #the index for the year, 2018:0, 2017:1, 2016:2...
    for row in range(1,102):
        if row % 6 == 1:
            week = row//6 + 1
            for team in team_results:
                team_results[team][year_i].append(team_results[team][year_i][week-1])  #making a result for the new week
        for col in range(4,20):
            cell = sheet.cell(row,col).value
            if cell == None:
                break
            if row % 6 == 1:                           #if the team beat the spread
                team_results[cell][year_i][week] += 1
            elif row % 6 == 5:                         #if the team lost to the spread
                team_results[cell][year_i][week] -= 1
            elif row % 6 == 2:
                if cell == 'home' and sheet.cell(row+1,col).value == 'favorite':
                    status_stats[year_i][week][0] += 1
                elif cell == 'away' and sheet.cell(row+1,col).value == 'underdog':
                    status_stats[year_i][week][1] += 1
                elif cell == 'home' and sheet.cell(row+1,col).value == 'underdog':
                    status_stats[year_i][week][2] += 1
                elif cell == 'away' and sheet.cell(row+1,col).value == 'favorite':
                    status_stats[year_i][week][3] += 1


#print(status_stats[0][1])

print(team_results['Baltimore Ravens'][0])
sh_overall = wk_analysis["Overall"]







wk_analysis.save(("prev_years_analysis.xlsx"))

"""


